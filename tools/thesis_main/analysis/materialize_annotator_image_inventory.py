"""Materialize a canonical P1-to-C2-A-RP worker/image inventory."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
STAGES = ("P1", "C1", "C2-B", "C2A-RP-B1", "C2A-RP-B2")
EXPECTED_STAGE_COUNTS = {"P1": 1481, "C1": 780, "C2-B": 160, "C2A-RP-B1": 40, "C2A-RP-B2": 40}
EXPECTED_CURRENT = {"1", "2", "6", "8", "10", "11", "12", "13", "15", "17", *map(str, range(28, 38))}
EXPECTED_WORKERS = EXPECTED_CURRENT | {"14", "18", "19", "21", "26", "27"}
EXPECTED_REPEATED_KEYS = {
    ("31", "wc2JMjhGNzB_9087f0358178420a8b9ac7b17a8919c7.jpg"),
    ("34", "wc2JMjhGNzB_55b45b0f19c2460bbcd1fb1c86c6610d.jpg"),
}
P1_IMPORTS = {
    "manual": ROOT / "import_json/stage1_prescreen_final_20260325/stage1_prescreen_manual_import_v2.json",
    "semi": ROOT / "import_json/stage1_prescreen_final_20260325/stage1_prescreen_semi_import_v5.json",
    "oos": ROOT / "import_json/stage1_prescreen_final_20260325/stage1_prescreen_oos_import_v2.json",
}
C1_ASSIGNMENT_EVIDENCE = ROOT / "analysis_results/c1_formal_audit_20260802_v16_final/c1_formal_audit_20260802_7fcacc5c2d6c_bf5def46_6bc67c03/c1_canonical_annotations.csv"
LATER_ASSIGNMENT_EVIDENCE = {
    "C2-B": ROOT / "analysis_results/c2b_closeout_20260806_final/c2b_canonical_submissions.csv",
    "C2A-RP-B1": ROOT / "analysis_results/c2a_rp_block1_reestimate_20260810_v1/c2a_rp_block1_canonical_submissions.csv",
    "C2A-RP-B2": ROOT / "analysis_results/c2a_rp_block2_reestimate_20260814_v1/c2a_rp_block2_canonical_submissions.csv",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def worker_id(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text.startswith("W"):
        text = text[1:]
    return str(int(text)) if text.isdigit() else text


def truth(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes"}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def workbook_rows(path: Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def load_identities(chinese_roster: Path, foreign_roster: Path) -> dict[str, dict[str, str]]:
    identities: dict[str, dict[str, str]] = {}
    for row in workbook_rows(chinese_roster)[1:]:
        wid = worker_id(row[0] if row else "")
        if not wid or not wid.isdigit() or len(row) < 3 or not row[2]:
            continue
        label = str(row[2]).strip()
        display = label if wid in {"19", "21"} else label.split()[-1]
        identities[wid] = {"display_name": display, "roster_name_text": label, "language_group": "Chinese"}
    for row in workbook_rows(foreign_roster)[1:]:
        wid = worker_id(row[0] if row else "")
        if not wid or not wid.isdigit() or len(row) < 2 or not row[1]:
            continue
        label = str(row[1]).strip()
        if wid in identities and identities[wid]["roster_name_text"] != label:
            raise ValueError(f"conflicting identity mapping for worker {wid}")
        identities[wid] = {"display_name": label, "roster_name_text": label, "language_group": "English"}
    return identities


def lifecycle_status(wid: str, current: set[str], profile: dict[str, str], prescreen: dict[str, str]) -> str:
    if wid in current:
        return "current_20"
    if profile.get("c2a_rp_completion_status") == "withdrawn":
        return "withdrawn_before_C2A_RP"
    if profile.get("completion_status") == "administrative_exclusion":
        return "administrative_exclusion"
    if truth(prescreen.get("exclude_from_primary_candidate")):
        return "P1_only_excluded"
    return "not_current"


def normalized_language(value: str) -> str:
    return {"zh": "Chinese", "en": "English"}.get(value.strip().lower(), value)


def materialize(
    *,
    raw_annotation_fact: Path,
    worker_fact: Path,
    current_roster: Path,
    prescreen_roster: Path,
    chinese_roster: Path,
    foreign_roster: Path,
    exit_roster: Path,
    output: Path,
) -> dict[str, int]:
    raw = read_csv(raw_annotation_fact)
    profiles = {worker_id(row["worker_id"]): row for row in read_csv(worker_fact)}
    prescreen = {worker_id(row["annotator_id"]): row for row in read_csv(prescreen_roster)}
    current_rows = read_csv(current_roster)
    current = {worker_id(row["worker_id"]) for row in current_rows}
    identities = load_identities(chinese_roster, foreign_roster)
    exit_names = {str(row[0]).strip() for row in workbook_rows(exit_roster)[1:] if row and row[0]}

    matched = [row for row in raw if row["canonical_join_status"] == "matched"]
    stage_counts = Counter(row["stage"] for row in matched)
    actual_workers = {worker_id(row["worker_id"]) for row in matched}
    canonical_ids = [row["canonical_annotation_id"] for row in matched]
    parsed: list[dict[str, Any]] = []
    missing_image_fields: list[str] = []
    image_identity_errors: list[str] = []

    for row in matched:
        data = json.loads(row["task_data_json"])
        wid = worker_id(row["worker_id"])
        image_title = str(data.get("title") or "").strip()
        image_url = str(data.get("image") or "").strip()
        base_task_id = str(row.get("base_task_id") or data.get("base_task_id") or "").strip()
        if not image_title or not image_url or not base_task_id:
            missing_image_fields.append(f"{wid}:{row['stage']}:{row['annotation_id']}")
        if Path(urlparse(image_url).path).name != image_title or image_title != f"{base_task_id}.jpg":
            image_identity_errors.append(f"{wid}:{row['stage']}:{row['annotation_id']}")
        parsed.append(
            {
                "worker_id": wid,
                "stage": row["stage"],
                "condition": row["condition"],
                "project_id": row["project_id"],
                "runtime_task_id": row["ls_runtime_task_id"],
                "task_data_task_id": row["task_data_task_id"],
                "base_task_id": base_task_id,
                "image_title": image_title,
                "image_url": image_url,
                "annotation_id": row["annotation_id"],
                "canonical_annotation_id": row["canonical_annotation_id"],
                "created_at": row["created_at"],
                "updated_at": row["updated_at"],
                "lead_time_seconds": row["lead_time_seconds"],
                "source_path": row["source_path"],
                "source_sha256": row["source_sha256"],
            }
        )

    p1_pools = {
        condition: {str(item["data"]["base_task_id"]) for item in json.loads(path.read_text(encoding="utf-8-sig"))}
        for condition, path in P1_IMPORTS.items()
    }
    c1_evidence = {row["canonical_annotation_id"]: row for row in read_csv(C1_ASSIGNMENT_EVIDENCE)}
    c1_assigned_worker_images = {
        (worker_id(row["worker_id"]), row["base_task_id"])
        for row in c1_evidence.values()
        if truth(row["assigned_expected"])
    }
    later_evidence = {
        stage: {
            (row["project_id"], row["runtime_task_id"], worker_id(row["worker_id"]), row["annotation_id"]): row
            for row in read_csv(path)
        }
        for stage, path in LATER_ASSIGNMENT_EVIDENCE.items()
    }
    for row in parsed:
        stage = row["stage"]
        canonical_id = row["canonical_annotation_id"]
        if stage == "P1":
            in_pool = row["base_task_id"] in p1_pools.get(row["condition"], set())
            row.update(
                assignment_match_status="stage_pool_member_worker_assignment_not_evaluable" if in_pool else "outside_stage_pool",
                assignment_provenance="P1_stage_pool_membership_only_no_worker_image_manifest",
                outside_assignment_subtype="",
                assigned_expected="",
                assignment_evidence_source=";".join(path.as_posix() for path in P1_IMPORTS.values()),
            )
        elif stage == "C1":
            evidence = c1_evidence.get(canonical_id)
            if not evidence:
                raise ValueError(f"missing C1 assignment evidence for {canonical_id}")
            provenance = evidence["assignment_provenance"]
            status = "outside_assignment_submission" if truth(evidence["outside_assignment_submission"]) else provenance
            subtype = ""
            if status == "outside_assignment_submission":
                subtype = (
                    "same_image_assigned_in_other_condition"
                    if (row["worker_id"], row["base_task_id"]) in c1_assigned_worker_images
                    else "image_not_assigned_to_worker"
                )
            row.update(
                assignment_match_status=status,
                assignment_provenance=provenance,
                outside_assignment_subtype=subtype,
                assigned_expected=evidence["assigned_expected"].lower(),
                assignment_evidence_source=C1_ASSIGNMENT_EVIDENCE.as_posix(),
            )
        else:
            evidence = later_evidence[stage].get((row["project_id"], row["runtime_task_id"], row["worker_id"], row["annotation_id"]))
            if not evidence:
                raise ValueError(f"missing {stage} assignment evidence for {canonical_id}")
            assigned = truth(evidence["formal_assignment_eligible"])
            row.update(
                assignment_match_status="formal_assignment" if assigned else "outside_assignment_submission",
                assignment_provenance="formal_assignment_eligible" if assigned else "formal_assignment_ineligible",
                outside_assignment_subtype="" if assigned else "unassigned_image",
                assigned_expected=str(assigned).lower(),
                assignment_evidence_source=LATER_ASSIGNMENT_EVIDENCE[stage].as_posix(),
            )

    worker_image_counts = Counter((row["worker_id"], row["image_title"]) for row in parsed)
    repeated_keys = {key for key, count in worker_image_counts.items() if count > 1}
    by_worker: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in parsed:
        by_worker[row["worker_id"]].append(row)

    matched_by_runtime_owner = defaultdict(list)
    for row in matched:
        matched_by_runtime_owner[(row["stage"], row["project_id"], row["ls_runtime_task_id"], worker_id(row["worker_id"]))].append(row)
    noncanonical: list[dict[str, Any]] = []
    for row in (item for item in raw if item["canonical_join_status"] != "matched"):
        key = (row["stage"], row["project_id"], row["ls_runtime_task_id"], worker_id(row["worker_id"]))
        selected = matched_by_runtime_owner[key]
        if len(selected) != 1:
            raise ValueError(f"noncanonical row does not resolve to one selected annotation: {key}")
        selected_row = selected[0]
        data = json.loads(row["task_data_json"])
        same_result = row["result_json"] == selected_row["result_json"]
        noncanonical.append(
            {
                "stage": row["stage"],
                "project_id": row["project_id"],
                "runtime_task_id": row["ls_runtime_task_id"],
                "worker_id": worker_id(row["worker_id"]),
                "image_title": data.get("title", ""),
                "base_task_id": row["base_task_id"],
                "excluded_annotation_id": row["annotation_id"],
                "selected_annotation_id": selected_row["annotation_id"],
                "selected_canonical_annotation_id": selected_row["canonical_annotation_id"],
                "same_result_json_as_selected": str(same_result).lower(),
                "relation_to_selected": "exact_result_duplicate" if same_result else "superseded_nonidentical_version",
                "canonical_join_status": row["canonical_join_status"],
                "canonical_join_key_rule": row["canonical_join_key_rule"],
                "excluded_created_at": row["created_at"],
                "excluded_updated_at": row["updated_at"],
                "source_path": row["source_path"],
                "source_sha256": row["source_sha256"],
            }
        )

    checks: list[dict[str, str]] = []

    def check(name: str, observed: Any, expected: Any, *, warning: bool = False, details: str = "") -> None:
        passed = observed == expected
        checks.append(
            {
                "check": name,
                "observed": str(observed),
                "expected": str(expected),
                "status": "warning" if warning and passed else "pass" if passed else "fail",
                "details": details,
            }
        )

    check("raw_annotation_count", len(raw), 2513)
    check("canonical_submission_count", len(matched), 2501)
    check("canonical_annotation_id_unique", len(set(canonical_ids)), 2501)
    check("stage_counts", dict(stage_counts), EXPECTED_STAGE_COUNTS)
    check("actual_worker_ids", sorted(actual_workers, key=int), sorted(EXPECTED_WORKERS, key=int))
    check("current_20_ids", sorted(current, key=int), sorted(EXPECTED_CURRENT, key=int))
    check("current_20_count", len(current), 20)
    check("missing_image_fields", missing_image_fields, [])
    check("image_title_url_base_identity", image_identity_errors, [])
    check("repeated_worker_image_keys", sorted(repeated_keys), sorted(EXPECTED_REPEATED_KEYS))
    check(
        "raw_noncanonical_by_stage",
        dict(Counter(row["stage"] for row in raw if row["canonical_join_status"] != "matched")),
        {"P1": 4, "C1": 8},
    )
    current_stage_errors = {
        wid: dict(Counter(row["stage"] for row in by_worker[wid]))
        for wid in current
        if set(row["stage"] for row in by_worker[wid]) != set(STAGES)
        or sum(row["stage"] == "C2A-RP-B2" for row in by_worker[wid]) != 2
    }
    check("current_20_stage_coverage", current_stage_errors, {})
    profile_completed = {wid for wid, row in profiles.items() if row.get("c2a_rp_completion_status") == "completed"}
    check("current_20_matches_completed_profile", sorted(current, key=int), sorted(profile_completed, key=int))
    unresolved = sorted(actual_workers - set(identities), key=int)
    check("identity_mapping_unresolved", unresolved, ["26"], warning=True, details="worker 26 has no name in current Chinese/foreign roster workbooks")
    check("identity_source_value_needs_confirmation", ["17:张fl"], ["17:张fl"], warning=True, details="preserved verbatim from Chinese roster; confirm real name manually if needed")
    check("noncanonical_detail_count", len(noncanonical), 12)
    unresolved_assignment = [row["canonical_annotation_id"] for row in parsed if not row["assignment_match_status"]]
    check("assignment_evidence_complete", unresolved_assignment, [])
    outside_by_worker = Counter(row["worker_id"] for row in parsed if row["assignment_match_status"] == "outside_assignment_submission")
    check("outside_assignment_by_worker", dict(outside_by_worker), {"10": 1, "31": 6, "34": 2})
    check("outside_assignment_non_C1", sum(row["stage"] != "C1" and row["assignment_match_status"] == "outside_assignment_submission" for row in parsed), 0)
    check("P1_outside_stage_pool", sum(row["assignment_match_status"] == "outside_stage_pool" for row in parsed), 0)
    check(
        "P1_worker_image_assignment_not_evaluable",
        sum(row["assignment_match_status"] == "stage_pool_member_worker_assignment_not_evaluable" for row in parsed),
        1481,
        warning=True,
        details="P1 has a fixed stage task pool and worker condition totals, but no worker-by-image assignment manifest",
    )
    check("C1_authorized_replacement_count", sum(row["assignment_match_status"] == "authorized_replacement_assignment" for row in parsed), 20)

    if any(row["status"] == "fail" for row in checks):
        failures = "; ".join(row["check"] for row in checks if row["status"] == "fail")
        raise ValueError(f"inventory reconciliation failed: {failures}")

    summary: list[dict[str, Any]] = []
    detail: list[dict[str, Any]] = []
    for wid in sorted(actual_workers, key=lambda value: (value not in current, int(value))):
        identity = identities.get(
            wid,
            {
                "display_name": "",
                "roster_name_text": "",
                "language_group": normalized_language(prescreen.get(wid, {}).get("language", "")),
            },
        )
        rows = by_worker[wid]
        exit_match = bool(identity["display_name"] and any(identity["display_name"] in name for name in exit_names))
        conflict = "historical_exit_form_but_later_submitted_through_C2A_RP_B2" if wid in current and exit_match else ""
        repeated_count = sum(worker == wid for worker, _ in repeated_keys)
        profile = profiles.get(wid, {})
        screening = prescreen.get(wid, {})
        status = lifecycle_status(wid, current, profile, screening)
        worker_summary: dict[str, Any] = {
            "worker_id": wid,
            "display_name": identity["display_name"],
            "roster_name_text": identity["roster_name_text"],
            "language_group": identity["language_group"],
            "current_20": str(wid in current).lower(),
            "lifecycle_status": status,
            "identity_mapping_status": "mapped" if wid in identities else "unresolved",
            "identity_note": "source_value_partial_or_pseudonymous" if wid == "17" else "",
            "historical_exit_form_match": str(exit_match).lower(),
            "status_conflict_note": conflict,
            "canonical_submission_count": len(rows),
            "unique_image_count": len({row["image_title"] for row in rows}),
            "repeated_worker_image_count": repeated_count,
            "outside_assignment_submission_count": sum(row["assignment_match_status"] == "outside_assignment_submission" for row in rows),
            "authorized_replacement_submission_count": sum(row["assignment_match_status"] == "authorized_replacement_assignment" for row in rows),
            "assignment_not_evaluable_submission_count": sum("not_evaluable" in row["assignment_match_status"] for row in rows),
            "first_created_at": min(row["created_at"] for row in rows),
            "last_updated_at": max(row["updated_at"] for row in rows),
            "profile_completion_status": profile.get("completion_status", ""),
            "c2a_rp_completion_status": profile.get("c2a_rp_completion_status", ""),
            "prescreen_exclude_from_primary_candidate": screening.get("exclude_from_primary_candidate", ""),
        }
        for stage in STAGES:
            key = stage.lower().replace("-", "_")
            stage_rows = [row for row in rows if row["stage"] == stage]
            worker_summary[f"{key}_submission_count"] = len(stage_rows)
            worker_summary[f"{key}_unique_image_count"] = len({row["image_title"] for row in stage_rows})
        summary.append(worker_summary)
        for row in rows:
            detail.append(
                {
                    "worker_id": wid,
                    "display_name": identity["display_name"],
                    "language_group": identity["language_group"],
                    "current_20": str(wid in current).lower(),
                    "lifecycle_status": status,
                    **row,
                    "worker_image_submission_count": worker_image_counts[(wid, row["image_title"])],
                    "worker_image_repeat": str(worker_image_counts[(wid, row["image_title"])] > 1).lower(),
                }
            )

    stage_rank = {stage: index for index, stage in enumerate(STAGES)}
    detail.sort(key=lambda row: (row["worker_id"] not in current, int(row["worker_id"]), stage_rank[row["stage"]], row["image_title"], int(row["annotation_id"])))
    summary_fields = list(summary[0])
    detail_fields = list(detail[0])
    check_fields = ["check", "observed", "expected", "status", "details"]
    noncanonical_fields = list(noncanonical[0])
    outside_assignment = [row for row in detail if row["assignment_match_status"] == "outside_assignment_submission"]
    for source in (
        raw_annotation_fact,
        worker_fact,
        current_roster,
        prescreen_roster,
        chinese_roster,
        foreign_roster,
        exit_roster,
        *P1_IMPORTS.values(),
        C1_ASSIGNMENT_EVIDENCE,
        *LATER_ASSIGNMENT_EVIDENCE.values(),
    ):
        checks.append({"check": f"source_sha256:{source.as_posix()}", "observed": sha256(source), "expected": "recorded", "status": "pass", "details": ""})

    output.mkdir(parents=True, exist_ok=True)
    write_csv(output / "annotator_summary.csv", summary, summary_fields)
    write_csv(output / "annotator_image_inventory.csv", detail, detail_fields)
    write_csv(output / "noncanonical_annotation_audit.csv", noncanonical, noncanonical_fields)
    write_csv(output / "outside_assignment_annotation_audit.csv", outside_assignment, detail_fields)
    write_csv(output / "inventory_checks.csv", checks, check_fields)
    return {"canonical_submissions": len(detail), "workers": len(summary), "current_workers": len(current)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw-annotation-fact", type=Path, required=True)
    parser.add_argument("--worker-fact", type=Path, required=True)
    parser.add_argument("--current-roster", type=Path, required=True)
    parser.add_argument("--prescreen-roster", type=Path, required=True)
    parser.add_argument("--chinese-roster", type=Path, required=True)
    parser.add_argument("--foreign-roster", type=Path, required=True)
    parser.add_argument("--exit-roster", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(materialize(**vars(args)), ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
