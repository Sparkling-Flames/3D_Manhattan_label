#!/usr/bin/env python3
"""Rebuild Stage 1 Chinese annotation completion workbook.

This is an operational helper for the current Stage 1 Chinese Label Studio
exports. It keeps the completion-rate contract separate from active-time
process audit sheets.
"""

from __future__ import annotations

import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[3]
EXPORT_DIR = ROOT / "export_label"
STAGE_DIR = EXPORT_DIR / "stage1_chinese"
OUT = STAGE_DIR / "标注完成情况完整结果.xlsx"
PROJECT_NAME = {28: "project-28", 29: "project-29", 30: "project-30"}
PROJECT_TYPE = {28: "manual", 29: "semi", 30: "oos"}
ADMIN_NONPARTICIPANT_IDS: set[str] = set()
FORCED_EXIT_IDS: set[str] = set()
ACCOUNTABLE_EXIT_OVERRIDE_IDS = {"11"}
FUTURE_ACCOUNTABLE_IDS = {"1", "2", "18"}


def worker_status(worker_id: str, exit_ids: set[str]) -> str:
    if worker_id in ACCOUNTABLE_EXIT_OVERRIDE_IDS:
        return "退出名单但暂计入"
    if worker_id in FUTURE_ACCOUNTABLE_IDS:
        return "后续参与待补齐"
    if worker_id in ADMIN_NONPARTICIPANT_IDS:
        return "管理人员暂不参与"
    if worker_id in exit_ids:
        return "退出人员"
    return "普通标注人员"


def is_accountable_worker(worker_id: str, exit_ids: set[str]) -> bool:
    return worker_id not in ADMIN_NONPARTICIPANT_IDS and worker_id not in exit_ids


def compact_name(value: object) -> str:
    text = str(value or "")
    return "".join(
        ch for ch in text if ("\u4e00" <= ch <= "\u9fff") or (ch.isascii() and ch.isalnum())
    ).lower()


def load_staff() -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    staff_path = next(EXPORT_DIR.glob("标注人员.xlsx"))
    wb = load_workbook(staff_path, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row) + [None] * 4
        number, group, name, available = vals[:4]
        if not any(v is not None for v in vals[:4]):
            continue
        if isinstance(number, str) and "红色" in number:
            continue
        if number is None and not name:
            continue
        if isinstance(number, str) and not number.strip().isdigit():
            continue
        rows.append(
            {
                "编号": int(number)
                if isinstance(number, (int, float)) and number == int(number)
                else (int(str(number).strip()) if str(number).strip().isdigit() else None),
                "组号": group,
                "人员": name,
                "可标注时间": available,
            }
        )
    return rows, {str(r["编号"]): r for r in rows if r["编号"] is not None}


def load_exit_mapping(staff_by_id: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], set[str]]:
    exit_path = next(EXPORT_DIR.glob("退出标注.xlsx"))
    wb = load_workbook(exit_path, data_only=True)
    ws = wb.active
    raw_rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row) + [None] * 9
        if not any(v is not None for v in vals[:9]):
            continue
        raw_rows.append(
            {
                "姓名": vals[0],
                "提交答卷时间": vals[1],
                "所用时间": vals[2],
                "来源": vals[3],
                "来源详情": vals[4],
                "来自IP": vals[5],
                "参与意愿": vals[6],
                "其他说明": vals[7],
                "退出原因": vals[8],
            }
        )

    mapped_rows: list[dict[str, Any]] = []
    exit_ids: set[str] = set()
    for raw_row in raw_rows:
        raw_name = str(raw_row.get("姓名") or "")
        compact = compact_name(raw_name)
        matched: str | None = None
        for staff_id, staff in staff_by_id.items():
            staff_name = str(staff.get("人员") or "")
            staff_compact = compact_name(staff_name)
            if compact and (compact in staff_compact or staff_compact in compact):
                matched = staff_id
                break
            for token in ["王皓", "周芷琪", "陈德昕"]:
                if token in raw_name and token in staff_name:
                    matched = staff_id
                    break
            if matched:
                break
        if matched:
            exit_ids.add(matched)
        mapped_rows.append(
            {
                **raw_row,
                "匹配编号": int(matched) if matched else None,
                "匹配人员": staff_by_id.get(matched, {}).get("人员") if matched else None,
            }
        )
    return mapped_rows, exit_ids


def load_project_data() -> dict[int, dict[str, Any]]:
    project_files = {
        28: sorted(STAGE_DIR.glob("project-28-*.json"))[-1],
        29: sorted(STAGE_DIR.glob("project-29-*.json"))[-1],
        30: sorted(STAGE_DIR.glob("project-30-*.json"))[-1],
    }
    result: dict[int, dict[str, Any]] = {}
    for project_id, path in project_files.items():
        tasks = json.loads(path.read_text(encoding="utf-8"))
        task_map: dict[str, dict[str, Any]] = {}
        worker_tasks: dict[str, set[str]] = defaultdict(set)
        valid_ann_count: dict[str, int] = defaultdict(int)
        cancel_count: dict[str, int] = defaultdict(int)
        for task in tasks:
            data = task.get("data") or {}
            task_key = str(data.get("task_id") or data.get("base_task_id") or task.get("id"))
            task_map[task_key] = {
                "LabelStudio任务ID": task.get("id"),
                "base_task_id": data.get("base_task_id"),
                "title": data.get("title"),
            }
            for ann in task.get("annotations") or []:
                worker_id = str(ann.get("completed_by")) if ann.get("completed_by") is not None else "unknown"
                if ann.get("was_cancelled"):
                    cancel_count[worker_id] += 1
                    continue
                worker_tasks[worker_id].add(task_key)
                valid_ann_count[worker_id] += 1
        result[project_id] = {
            "path": path,
            "task_ids": list(task_map.keys()),
            "task_map": task_map,
            "worker_tasks": worker_tasks,
            "worker_valid_ann_count": valid_ann_count,
            "worker_cancel_count": cancel_count,
        }
    return result


def build_active_time_rows(
    project_data: dict[int, dict[str, Any]],
    staff_by_id: dict[str, dict[str, Any]],
    exit_ids: set[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    active_dir = ROOT / "active_logs" / "new_server"
    groups: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "event_count": 0,
            "unique_tasks": set(),
            "active_seconds_fragment_sum": 0.0,
            "active_seconds_max": 0.0,
            "first_file": None,
            "last_file": None,
            "project_names": set(),
            "script_versions": set(),
        }
    )
    parse_error_count = 0
    for path in sorted(active_dir.glob("active_times_*.jsonl")):
        with path.open("r", encoding="utf-8", errors="replace") as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    item = json.loads(line)
                except Exception:
                    parse_error_count += 1
                    continue
                project_id = str(item.get("project_id") or "")
                if project_id not in {"28", "29", "30"}:
                    continue
                annotator_id = item.get("annotator_id")
                annotator_id = (
                    str(annotator_id).strip()
                    if annotator_id is not None and str(annotator_id).strip()
                    else "unknown"
                )
                task_id = str(item.get("task_id") or "").strip() or "unknown"
                group = groups[(project_id, annotator_id)]
                group["event_count"] += 1
                group["unique_tasks"].add(task_id)
                group["active_seconds_fragment_sum"] += float(item.get("active_seconds_fragment") or 0)
                group["active_seconds_max"] = max(
                    group["active_seconds_max"], float(item.get("active_seconds") or 0)
                )
                group["first_file"] = group["first_file"] or path.name
                group["last_file"] = path.name
                if item.get("project_name"):
                    group["project_names"].add(str(item.get("project_name")))
                if item.get("script_version"):
                    group["script_versions"].add(str(item.get("script_version")))

    summary_rows: list[dict[str, Any]] = []
    anomaly_rows: list[dict[str, Any]] = []
    for (project_id_s, annotator_id), group in sorted(
        groups.items(), key=lambda kv: (int(kv[0][0]), int(kv[0][1]) if kv[0][1].isdigit() else 9999, kv[0][1])
    ):
        project_id = int(project_id_s)
        staff = staff_by_id.get(annotator_id)
        in_staff = annotator_id in staff_by_id
        is_exited = annotator_id in exit_ids
        status = worker_status(annotator_id, exit_ids)
        appeared_in_json = annotator_id in project_data[project_id]["worker_tasks"]
        anomaly_types: list[str] = []
        if annotator_id in {"", "unknown", "None", "null"} or not annotator_id.isdigit():
            anomaly_types.append("active_time_annotator_id_missing_or_non_numeric")
        elif not in_staff:
            anomaly_types.append("active_time_annotator_id_not_in_staff_excel")
        if annotator_id in ADMIN_NONPARTICIPANT_IDS:
            anomaly_types.append("active_time_from_admin_nonparticipant")
        if is_exited:
            anomaly_types.append("active_time_from_exited_person")
        if not appeared_in_json:
            anomaly_types.append("active_time_without_valid_annotation_in_project_json")
        row = {
            "项目ID": project_id,
            "项目": PROJECT_NAME[project_id],
            "项目类型": PROJECT_TYPE[project_id],
            "active_annotator_id": annotator_id,
            "人员表匹配": "是" if in_staff else "否",
            "人员": staff.get("人员") if staff else None,
            "人员状态": status,
            "退出人员": "是" if is_exited else "否",
            "JSON中是否有有效标注": "是" if appeared_in_json else "否",
            "JSON有效完成任务数": len(project_data[project_id]["worker_tasks"].get(annotator_id, set())),
            "active_log事件数": group["event_count"],
            "active_log唯一task数": len(group["unique_tasks"]),
            "active_seconds_fragment_sum": round(group["active_seconds_fragment_sum"], 3),
            "active_seconds_max": round(group["active_seconds_max"], 3),
            "project_name日志值": "；".join(sorted(group["project_names"])),
            "script_versions": "；".join(sorted(group["script_versions"])),
            "first_log_file": group["first_file"],
            "last_log_file": group["last_file"],
            "异常类型": "；".join(anomaly_types) if anomaly_types else None,
        }
        summary_rows.append(row)
        if anomaly_types:
            anomaly_rows.append(row)
    return summary_rows, anomaly_rows, parse_error_count


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], headers: list[str] | None = None):
    ws = wb.create_sheet(title)
    if headers is None:
        headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="DDDDDD")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = max((len(str(cell.value)) for cell in ws[letter] if cell.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)
    return ws


def main() -> None:
    staff_rows, staff_by_id = load_staff()
    numbered_staff = [r for r in staff_rows if r["编号"] is not None]
    exit_rows, exit_ids = load_exit_mapping(staff_by_id)
    exit_ids = (set(exit_ids) | FORCED_EXIT_IDS) - ACCOUNTABLE_EXIT_OVERRIDE_IDS
    project_data = load_project_data()
    active_summary_rows, active_anomaly_rows, parse_error_count = build_active_time_rows(
        project_data, staff_by_id, exit_ids
    )

    project_relation_rows = [
        {
            "项目ID": project_id,
            "项目": PROJECT_NAME[project_id],
            "项目类型": PROJECT_TYPE[project_id],
            "说明": {
                28: "P1 manual 池；纯人工几何能力基准。",
                29: "P1 semi 池；半自动初始化 / trap 与 control。",
                30: "P1 oos 池；scope gate / OOS 判定。",
            }[project_id],
        }
        for project_id in [28, 29, 30]
    ]
    summary_rows: list[dict[str, Any]] = []
    participant_detail: list[dict[str, Any]] = []
    incomplete_rows: list[dict[str, Any]] = []
    nonaccountable_rows: list[dict[str, Any]] = []
    missing_detail: list[dict[str, Any]] = []
    all_excel_rows: list[dict[str, Any]] = []
    p1_suite_rows: list[dict[str, Any]] = []

    any_participant_ids = set().union(*(set(project_data[pid]["worker_tasks"]) for pid in [28, 29, 30]))
    p1_candidate_ids = any_participant_ids | FUTURE_ACCOUNTABLE_IDS
    p1_accountable_ids = sorted(
        [worker_id for worker_id in p1_candidate_ids if is_accountable_worker(worker_id, exit_ids)],
        key=lambda x: int(x) if x.isdigit() else 9999,
    )
    p1_full_suite_complete_ids: list[str] = []
    p1_full_suite_incomplete: list[tuple[str, dict[int, list[str]]]] = []
    for worker_id in p1_accountable_ids:
        missing_by_project: dict[int, list[str]] = {}
        for project_id in [28, 29, 30]:
            data = project_data[project_id]
            done_set = data["worker_tasks"].get(worker_id, set())
            missing = [tid for tid in data["task_ids"] if tid not in done_set]
            missing_by_project[project_id] = missing
        if all(not missing_by_project[project_id] for project_id in [28, 29, 30]):
            p1_full_suite_complete_ids.append(worker_id)
        else:
            p1_full_suite_incomplete.append((worker_id, missing_by_project))
        staff = staff_by_id.get(worker_id, {})
        suite_row: dict[str, Any] = {
            "编号": int(worker_id) if worker_id.isdigit() else worker_id,
            "人员": staff.get("人员"),
            "人员状态": worker_status(worker_id, exit_ids),
            "纳入完成率责任口径": "是",
            "P1全套完成": "是" if all(not missing_by_project[pid] for pid in [28, 29, 30]) else "否",
        }
        for project_id in [28, 29, 30]:
            data = project_data[project_id]
            total = len(data["task_ids"])
            done = total - len(missing_by_project[project_id])
            suite_row[f"{PROJECT_TYPE[project_id]}完成"] = f"{done}/{total}"
            suite_row[f"{PROJECT_TYPE[project_id]}缺失数"] = len(missing_by_project[project_id])
            suite_row[f"{PROJECT_TYPE[project_id]}缺失任务"] = (
                "、".join(missing_by_project[project_id]) if missing_by_project[project_id] else None
            )
        p1_suite_rows.append(suite_row)

    for project_id in [28, 29, 30]:
        data = project_data[project_id]
        total = len(data["task_ids"])
        actual_ids = sorted(data["worker_tasks"], key=lambda x: int(x) if x.isdigit() else 9999)
        complete: list[str] = []
        incomplete: list[tuple[str, int, list[str]]] = []
        accountable_ids = [worker_id for worker_id in actual_ids if is_accountable_worker(worker_id, exit_ids)]
        accountable_complete: list[str] = []
        accountable_incomplete: list[tuple[str, int, list[str]]] = []
        for worker_id in actual_ids:
            done = len(data["worker_tasks"][worker_id])
            missing = [tid for tid in data["task_ids"] if tid not in data["worker_tasks"][worker_id]]
            staff = staff_by_id.get(worker_id, {})
            status_label = worker_status(worker_id, exit_ids)
            accountable = is_accountable_worker(worker_id, exit_ids)
            status = "完整" if done == total else "未完整"
            participant_detail.append(
                {
                    "项目ID": project_id,
                    "项目": PROJECT_NAME[project_id],
                    "项目类型": PROJECT_TYPE[project_id],
                    "编号": int(worker_id) if worker_id.isdigit() else worker_id,
                    "人员": staff.get("人员"),
                    "人员状态": status_label,
                    "纳入完成率责任口径": "是" if accountable else "否",
                    "是否实际参与": "是",
                    "完成状态": status,
                    "任务总数": total,
                    "已完成任务数(去重)": done,
                    "有效标注条数": data["worker_valid_ann_count"].get(worker_id, 0),
                    "缺失任务数": len(missing),
                    "取消标注任务数": data["worker_cancel_count"].get(worker_id, 0),
                    "缺失任务ID列表": "、".join(missing) if missing else None,
                    "是否退出人员": "是" if worker_id in exit_ids else "否",
                }
            )
            if done == total:
                complete.append(worker_id)
            else:
                incomplete.append((worker_id, done, missing))
            if accountable and done == total:
                accountable_complete.append(worker_id)
            elif accountable:
                accountable_incomplete.append((worker_id, done, missing))
            else:
                nonaccountable_rows.append(
                    {
                        "项目ID": project_id,
                        "项目": PROJECT_NAME[project_id],
                        "项目类型": PROJECT_TYPE[project_id],
                        "编号": int(worker_id) if worker_id.isdigit() else worker_id,
                        "人员": staff.get("人员"),
                        "人员状态": status_label,
                        "任务总数": total,
                        "已完成任务数(去重)": done,
                        "缺失任务数": len(missing),
                        "缺失任务ID列表": "、".join(missing) if missing else None,
                        "保留原因": "保留运行痕迹，但不纳入完成率责任口径",
                    }
                )
            if accountable and done != total:
                incomplete_rows.append(
                    {
                        "项目ID": project_id,
                        "项目": PROJECT_NAME[project_id],
                        "项目类型": PROJECT_TYPE[project_id],
                        "编号": int(worker_id) if worker_id.isdigit() else worker_id,
                        "人员": staff.get("人员"),
                        "人员状态": status_label,
                        "任务总数": total,
                        "已完成任务数(去重)": done,
                        "缺失任务数": len(missing),
                        "缺失任务ID列表": "、".join(missing),
                        "是否退出人员": "是" if worker_id in exit_ids else "否",
                    }
                )
                for task_id in missing:
                    task = data["task_map"].get(task_id, {})
                    missing_detail.append(
                        {
                            "项目ID": project_id,
                            "项目": PROJECT_NAME[project_id],
                            "项目类型": PROJECT_TYPE[project_id],
                            "编号": int(worker_id) if worker_id.isdigit() else worker_id,
                            "人员": staff.get("人员"),
                            "人员状态": status_label,
                            "缺失任务ID": task_id,
                            "LabelStudio任务ID": task.get("LabelStudio任务ID"),
                            "base_task_id": task.get("base_task_id"),
                            "title": task.get("title"),
                            "口径": "该项目JSON中实际参与者未覆盖全部任务",
                            "是否退出人员": "是" if worker_id in exit_ids else "否",
                        }
                    )
        miss_text = (
            "无"
            if not incomplete
            else "、".join(
                f"{worker_id} {staff_by_id.get(worker_id, {}).get('人员') or ''}({done}/{total})"
                for worker_id, done, _ in incomplete
            )
        )
        accountable_miss_text = (
            "无"
            if not accountable_incomplete
            else "、".join(
                f"{worker_id} {staff_by_id.get(worker_id, {}).get('人员') or ''}({done}/{total})"
                for worker_id, done, _ in accountable_incomplete
            )
        )
        summary_rows.append(
            {
                "项目ID": project_id,
                "项目": PROJECT_NAME[project_id],
                "项目类型": PROJECT_TYPE[project_id],
                "任务总数": total,
                "实际参与人数": len(actual_ids),
                "实际完整人数": len(complete),
                "实际未完整人数": len(incomplete),
                "实际未完整人员": miss_text,
                "责任口径参与人数": len(p1_accountable_ids),
                "责任口径完整人数": len(p1_full_suite_complete_ids),
                "责任口径未完整人数": len(p1_full_suite_incomplete),
                "责任口径未完整人员": "无"
                if not p1_full_suite_incomplete
                else "、".join(
                    f"{worker_id} {staff_by_id.get(worker_id, {}).get('人员') or ''}"
                    for worker_id, _ in p1_full_suite_incomplete
                ),
                "责任口径说明": "P1全套责任口径：责任人员必须完成 manual/semi/oos 三个项目",
            }
        )
        for staff in numbered_staff:
            worker_id = str(staff["编号"])
            done_set = data["worker_tasks"].get(worker_id, set())
            missing = [tid for tid in data["task_ids"] if tid not in done_set]
            all_excel_rows.append(
                {
                    "项目ID": project_id,
                    "项目": PROJECT_NAME[project_id],
                    "项目类型": PROJECT_TYPE[project_id],
                    "编号": staff["编号"],
                    "人员": staff["人员"],
                    "人员状态": worker_status(worker_id, exit_ids),
                    "纳入完成率责任口径": "是" if is_accountable_worker(worker_id, exit_ids) else "否",
                    "JSON中是否出现": "是" if worker_id in data["worker_tasks"] else "否",
                    "任务总数": total,
                    "已完成任务数(去重)": len(done_set),
                    "缺失任务数": len(missing),
                    "缺失任务ID列表": "、".join(missing) if missing else None,
                    "是否退出人员": "是" if worker_id in exit_ids else "否",
                }
            )

    incomplete_rows = []
    missing_detail = []
    for worker_id, missing_by_project in p1_full_suite_incomplete:
        staff = staff_by_id.get(worker_id, {})
        for project_id in [28, 29, 30]:
            data = project_data[project_id]
            total = len(data["task_ids"])
            missing = missing_by_project[project_id]
            done = total - len(missing)
            if not missing:
                continue
            incomplete_rows.append(
                {
                    "项目ID": project_id,
                    "项目": PROJECT_NAME[project_id],
                    "项目类型": PROJECT_TYPE[project_id],
                    "编号": int(worker_id) if worker_id.isdigit() else worker_id,
                    "人员": staff.get("人员"),
                    "人员状态": worker_status(worker_id, exit_ids),
                    "任务总数": total,
                    "已完成任务数(去重)": done,
                    "缺失任务数": len(missing),
                    "缺失任务ID列表": "、".join(missing),
                    "是否退出人员": "是" if worker_id in exit_ids else "否",
                    "口径": "P1全套责任口径：责任人员必须完成 manual/semi/oos 三个项目",
                }
            )
            for task_id in missing:
                task = data["task_map"].get(task_id, {})
                missing_detail.append(
                    {
                        "项目ID": project_id,
                        "项目": PROJECT_NAME[project_id],
                        "项目类型": PROJECT_TYPE[project_id],
                        "编号": int(worker_id) if worker_id.isdigit() else worker_id,
                        "人员": staff.get("人员"),
                        "人员状态": worker_status(worker_id, exit_ids),
                        "缺失任务ID": task_id,
                        "LabelStudio任务ID": task.get("LabelStudio任务ID"),
                        "base_task_id": task.get("base_task_id"),
                        "title": task.get("title"),
                        "口径": "P1全套责任口径：责任人员必须完成 manual/semi/oos 三个项目",
                        "是否退出人员": "是" if worker_id in exit_ids else "否",
                    }
                )

    person_rows: list[dict[str, Any]] = []
    for staff in numbered_staff:
        worker_id = str(staff["编号"])
        row = {
            "编号": staff["编号"],
            "人员": staff["人员"],
            "人员状态": worker_status(worker_id, exit_ids),
            "纳入完成率责任口径": "是" if is_accountable_worker(worker_id, exit_ids) else "否",
        }
        for project_id in [28, 29, 30]:
            data = project_data[project_id]
            total = len(data["task_ids"])
            done_set = data["worker_tasks"].get(worker_id, set())
            missing = [tid for tid in data["task_ids"] if tid not in done_set]
            row[f"{PROJECT_NAME[project_id]}类型"] = PROJECT_TYPE[project_id]
            row[f"{PROJECT_NAME[project_id]}完成"] = f"{len(done_set)}/{total}"
            row[f"{PROJECT_NAME[project_id]}缺失"] = len(missing)
            row[f"{PROJECT_NAME[project_id]}缺失任务"] = "、".join(missing) if missing else None
        row["是否退出人员"] = "是" if worker_id in exit_ids else "否"
        person_rows.append(row)

    backup = STAGE_DIR / f"标注完成情况完整结果.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if OUT.exists():
        shutil.copy2(OUT, backup)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "项目对应关系", project_relation_rows)
    add_sheet(wb, "汇总", summary_rows)
    add_sheet(wb, "P1全套责任口径", p1_suite_rows)
    add_sheet(wb, "人员一一对应", person_rows)
    add_sheet(wb, "未完整人员", incomplete_rows)
    add_sheet(wb, "非责任口径人员", nonaccountable_rows)
    add_sheet(wb, "实际参与者明细", participant_detail)
    add_sheet(wb, "缺失任务明细", missing_detail)
    add_sheet(wb, "所有Excel编号口径", all_excel_rows)
    add_sheet(wb, "Excel人员名单", staff_rows, headers=["编号", "组号", "人员", "可标注时间"])
    add_sheet(wb, "active_time异常", active_anomaly_rows)
    add_sheet(wb, "active_time汇总", active_summary_rows)
    add_sheet(
        wb,
        "退出人员映射",
        exit_rows,
        headers=[
            "姓名",
            "匹配编号",
            "匹配人员",
            "提交答卷时间",
            "所用时间",
            "来源",
            "来源详情",
            "来自IP",
            "参与意愿",
            "其他说明",
            "退出原因",
        ],
    )
    add_sheet(
        wb,
        "说明",
        [
            {
                "说明项": "对应关系",
                "内容": "Excel列“编号”对应 Label Studio JSON annotation.completed_by，也对应 active_logs 中 annotator_id。",
            },
            {
                "说明项": "项目类型",
                "内容": "project-28=manual，project-29=semi，project-30=oos；详见“项目对应关系”sheet。",
            },
            {
                "说明项": "主要口径",
                "内容": "“未完整人员”和汇总中的责任口径采用 P1 全套责任口径；责任人员必须同时完成 manual/semi/oos 三个项目，只完成单个项目不算完整完成。",
            },
            {
                "说明项": "辅助口径",
                "内容": "“所有Excel编号口径”把新版标注人员Excel中每个有编号的人都放到每个项目下统计，JSON中未出现的人会显示为未出现。",
            },
            {
                "说明项": "特殊人员口径",
                "内容": "1号、2号、18号后续会参与，纳入 P1 全套责任口径并标为“后续参与待补齐”；11号虽在退出名单映射中，但因实际完整参与，暂时纳入完成率责任口径。历史标注和 active_time 均不删除。",
            },
            {"说明项": "有效标注", "内容": "排除 was_cancelled=True 的标注。"},
            {
                "说明项": "active_time异常",
                "内容": "只作为过程审计；列出 project-28/29/30 中 active_time annotator_id 不在人员表、缺失/非数字、属于退出人员、管理人员暂不参与，或该项目JSON没有有效标注的日志主体。1号、2号、18号当前不因后续参与状态本身记异常。",
            },
            {
                "说明项": "active_time来源",
                "内容": "读取 active_logs/new_server 顶层 active_times_*.jsonl；不递归读取 legacy 子目录。",
            },
            {
                "说明项": "输入JSON",
                "内容": "使用 export_label/stage1_chinese 下 2026-06-10 的 project-28/29/30 导出 JSON。",
            },
            {"说明项": "active_time解析错误行数", "内容": parse_error_count},
            {"说明项": "备份", "内容": f"覆盖前备份为 {backup.name}"},
        ],
    )

    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="F4CCCC")
    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]
        for row_cells in ws.iter_rows(min_row=2):
            values = {headers[i]: row_cells[i].value for i in range(len(headers))}
            fill = None
            if values.get("异常类型"):
                fill = bad_fill
            elif (
                values.get("是否退出人员") == "是"
                or values.get("退出人员") == "是"
                or values.get("人员状态") in {"管理人员暂不参与", "退出人员"}
            ):
                fill = warn_fill
            if fill:
                for cell in row_cells:
                    cell.fill = fill

    wb.save(OUT)
    print(f"wrote: {OUT}")
    print(f"backup: {backup}")
    print("summary:")
    for row in summary_rows:
        print(row)
    print(f"active_time_anomaly_count: {len(active_anomaly_rows)}")
    for row in active_anomaly_rows:
        print(row)


if __name__ == "__main__":
    main()
